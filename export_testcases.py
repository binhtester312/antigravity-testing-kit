"""
Script xuất bộ Test Case Module Login ra file Excel (.xlsx)
Hệ thống: Perfex CRM - https://crm.anhtester.com/admin/authentication
"""

import sys
try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("📦 Đang cài openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

from datetime import datetime

# ─────────────────────────────────────────────
# DỮ LIỆU TEST CASE
# ─────────────────────────────────────────────
TESTCASES = [
    # ── MODULE I: ĐĂNG NHẬP (HIGH RISK) ──────────────────────────────────────────
    {
        "tc_id": "CRM_LOGIN_TC_001",
        "module": "Đăng nhập",
        "risk_level": "High",
        "test_scenario": "Đăng nhập thành công với email và password hợp lệ",
        "pre_condition": "Tài khoản tồn tại; Chưa đăng nhập; Đang ở trang /admin/authentication",
        "test_steps": (
            "1. Mở URL: https://crm.anhtester.com/admin/authentication\n"
            "2. Nhập Email vào trường Email Address\n"
            "3. Nhập Password vào trường Password\n"
            "4. Click nút 'Login'"
        ),
        "test_data": "Email: admin@example.com\nPassword: 12345678",
        "expected_result": (
            "- Chuyển hướng thành công đến https://crm.anhtester.com/admin/\n"
            "- Trang Dashboard hiển thị tiêu đề 'Dashboard'\n"
            "- Không có thông báo lỗi"
        ),
        "priority": "High",
    },
    {
        "tc_id": "CRM_LOGIN_TC_002",
        "module": "Đăng nhập",
        "risk_level": "High",
        "test_scenario": "Đăng nhập thành công + tick 'Remember me' — phiên duy trì sau khi đóng browser",
        "pre_condition": "Tài khoản tồn tại; Chưa đăng nhập",
        "test_steps": (
            "1. Mở trang Login\n"
            "2. Nhập Email\n"
            "3. Nhập Password\n"
            "4. Tick vào checkbox 'Remember me'\n"
            "5. Click 'Login'\n"
            "6. Đóng trình duyệt hoàn toàn\n"
            "7. Mở lại trình duyệt, truy cập https://crm.anhtester.com/admin/"
        ),
        "test_data": "Email: admin@example.com\nPassword: 12345678\nRemember me: ✅ checked",
        "expected_result": (
            "- Sau bước 5: Chuyển hướng đến Dashboard thành công\n"
            "- Sau bước 7: Vẫn ở trạng thái đã đăng nhập (không bị đẩy về Login)\n"
            "[A-04] Giả định: phiên duy trì 30 ngày"
        ),
        "priority": "High",
    },
    {
        "tc_id": "CRM_LOGIN_TC_003",
        "module": "Đăng nhập",
        "risk_level": "High",
        "test_scenario": "Để trống cả Email và Password — hiển thị 2 thông báo lỗi đồng thời (Decision Table)",
        "pre_condition": "Đang ở trang Login; Cả hai trường đang trống",
        "test_steps": (
            "1. Mở trang Login\n"
            "2. KHÔNG nhập gì vào Email\n"
            "3. KHÔNG nhập gì vào Password\n"
            "4. Click nút 'Login'"
        ),
        "test_data": "Email: (để trống)\nPassword: (để trống)",
        "expected_result": (
            "- Không chuyển hướng, vẫn ở trang Login\n"
            "- Hiển thị đồng thời 2 thông báo trong <div class='alert alert-danger text-center'>:\n"
            "  • 'The Email Address field is required.'\n"
            "  • 'The Password field is required.'"
        ),
        "priority": "High",
    },
    {
        "tc_id": "CRM_LOGIN_TC_004",
        "module": "Đăng nhập",
        "risk_level": "High",
        "test_scenario": "Để trống Email, nhập Password bất kỳ — lỗi email bắt buộc (Decision Table)",
        "pre_condition": "Đang ở trang Login",
        "test_steps": (
            "1. Mở trang Login\n"
            "2. KHÔNG nhập Email\n"
            "3. Nhập Password\n"
            "4. Click 'Login'"
        ),
        "test_data": "Email: (để trống)\nPassword: anyPassword99",
        "expected_result": (
            "- Không chuyển hướng\n"
            "- Hiển thị: 'The Email Address field is required.'\n"
            "- Không hiện lỗi về Password"
        ),
        "priority": "High",
    },
    {
        "tc_id": "CRM_LOGIN_TC_005",
        "module": "Đăng nhập",
        "risk_level": "High",
        "test_scenario": "Nhập Email hợp lệ, để trống Password — lỗi password bắt buộc (Decision Table)",
        "pre_condition": "Đang ở trang Login",
        "test_steps": (
            "1. Mở trang Login\n"
            "2. Nhập Email hợp lệ\n"
            "3. KHÔNG nhập Password\n"
            "4. Click 'Login'"
        ),
        "test_data": "Email: admin@example.com\nPassword: (để trống)",
        "expected_result": (
            "- Không chuyển hướng\n"
            "- Hiển thị: 'The Password field is required.'\n"
            "- Không hiện lỗi về Email"
        ),
        "priority": "High",
    },
    {
        "tc_id": "CRM_LOGIN_TC_006",
        "module": "Đăng nhập",
        "risk_level": "High",
        "test_scenario": "Email sai định dạng — thiếu ký tự '@' → client-side HTML5 validation (EP)",
        "pre_condition": "Đang ở trang Login",
        "test_steps": (
            "1. Mở trang Login\n"
            "2. Nhập Email không có ký tự '@'\n"
            "3. Nhập Password bất kỳ\n"
            "4. Click 'Login'"
        ),
        "test_data": "Email: invalidemail\nPassword: anyPassword99",
        "expected_result": (
            "- Trình duyệt chặn form, không submit\n"
            "- Hiển thị HTML5 tooltip: 'Please include an '@' in the email address.'\n"
            "- Không có request server-side"
        ),
        "priority": "High",
    },
    {
        "tc_id": "CRM_LOGIN_TC_007",
        "module": "Đăng nhập",
        "risk_level": "Medium",
        "test_scenario": "Email có '@' nhưng thiếu domain — BVA biên định dạng email",
        "pre_condition": "Đang ở trang Login",
        "test_steps": (
            "1. Mở trang Login\n"
            "2. Nhập Email có '@' nhưng không có domain\n"
            "3. Nhập Password bất kỳ\n"
            "4. Click 'Login'"
        ),
        "test_data": "Email: user@\nPassword: anyPassword99",
        "expected_result": (
            "- Trình duyệt chặn form, không submit\n"
            "- Hiển thị HTML5 tooltip yêu cầu nhập phần domain sau '@'"
        ),
        "priority": "Medium",
    },
    {
        "tc_id": "CRM_LOGIN_TC_008",
        "module": "Đăng nhập",
        "risk_level": "High",
        "test_scenario": "Email đúng định dạng nhưng không tồn tại — lỗi server-side chung (EP)",
        "pre_condition": "Đang ở trang Login",
        "test_steps": (
            "1. Mở trang Login\n"
            "2. Nhập Email đúng định dạng nhưng không tồn tại trong hệ thống\n"
            "3. Nhập Password bất kỳ\n"
            "4. Click 'Login'"
        ),
        "test_data": "Email: wrong_user_99@gmail.com\nPassword: AnyPass@123",
        "expected_result": (
            "- Không chuyển hướng\n"
            "- Hiển thị (server-side): 'Invalid email or password'\n"
            "- Thông báo chung, KHÔNG tiết lộ email có tồn tại hay không (NFR-03)"
        ),
        "priority": "High",
    },
    {
        "tc_id": "CRM_LOGIN_TC_009",
        "module": "Đăng nhập",
        "risk_level": "High",
        "test_scenario": "Email đúng + Password sai — lỗi server-side (EP)",
        "pre_condition": "Đang ở trang Login",
        "test_steps": (
            "1. Mở trang Login\n"
            "2. Nhập Email đúng của tài khoản có thật\n"
            "3. Nhập Password SAI\n"
            "4. Click 'Login'"
        ),
        "test_data": "Email: admin@example.com\nPassword: WrongPass@9999",
        "expected_result": (
            "- Không chuyển hướng\n"
            "- Hiển thị: 'Invalid email or password'\n"
            "- Thông báo chung, không tiết lộ password sai"
        ),
        "priority": "High",
    },
    {
        "tc_id": "CRM_LOGIN_TC_010",
        "module": "Đăng nhập",
        "risk_level": "Medium",
        "test_scenario": "Submit form bằng phím Enter thay vì click nút Login (Alternate Path)",
        "pre_condition": "Đang ở trang Login",
        "test_steps": (
            "1. Mở trang Login\n"
            "2. Nhập Email hợp lệ\n"
            "3. Nhập Password đúng\n"
            "4. Nhấn phím Enter (không click nút)"
        ),
        "test_data": "Email: admin@example.com\nPassword: 12345678",
        "expected_result": (
            "- Form được submit\n"
            "- Chuyển hướng đến Dashboard thành công\n"
            "[A-08] Giả định: form hỗ trợ Enter"
        ),
        "priority": "Medium",
    },
    {
        "tc_id": "CRM_LOGIN_TC_011",
        "module": "Đăng nhập",
        "risk_level": "Medium",
        "test_scenario": "Email có khoảng trắng đầu/cuối (leading/trailing spaces) — BVA edge case",
        "pre_condition": "Đang ở trang Login",
        "test_steps": (
            "1. Mở trang Login\n"
            "2. Nhập Email có 2 dấu cách ở đầu và 2 dấu cách ở cuối\n"
            "3. Nhập Password đúng\n"
            "4. Click 'Login'"
        ),
        "test_data": "Email: '  admin@example.com  ' (2 spaces đầu, 2 spaces cuối)\nPassword: 12345678",
        "expected_result": (
            "[A-03] Nếu server tự trim: Đăng nhập thành công → redirect Dashboard\n"
            "Nếu server KHÔNG trim: Hiển thị 'Invalid email or password'"
        ),
        "priority": "Medium",
    },
    {
        "tc_id": "CRM_LOGIN_TC_012",
        "module": "Đăng nhập",
        "risk_level": "Medium",
        "test_scenario": "Password 1 ký tự — BVA biên dưới",
        "pre_condition": "Đang ở trang Login",
        "test_steps": (
            "1. Mở trang Login\n"
            "2. Nhập Email đúng\n"
            "3. Nhập Password chỉ 1 ký tự\n"
            "4. Click 'Login'"
        ),
        "test_data": "Email: admin@example.com\nPassword: A (1 ký tự)",
        "expected_result": (
            "- Không hiện lỗi client-side về độ dài\n"
            "- Form submit lên server\n"
            "- Kết quả: 'Invalid email or password'\n"
            "[A-02] Giả định: không có min/max length"
        ),
        "priority": "Medium",
    },
    {
        "tc_id": "CRM_LOGIN_TC_013",
        "module": "Đăng nhập",
        "risk_level": "Low",
        "test_scenario": "Password 255 ký tự — BVA biên trên",
        "pre_condition": "Đang ở trang Login",
        "test_steps": (
            "1. Mở trang Login\n"
            "2. Nhập Email đúng\n"
            "3. Nhập Password dài 255 ký tự\n"
            "4. Click 'Login'"
        ),
        "test_data": "Email: admin@example.com\nPassword: Aa1! lặp lại đủ 255 ký tự",
        "expected_result": (
            "- Form không crash, không timeout\n"
            "- Không có lỗi client-side\n"
            "- Server trả về 'Invalid email or password'\n"
            "[A-02]"
        ),
        "priority": "Low",
    },
    {
        "tc_id": "CRM_LOGIN_TC_014",
        "module": "Đăng nhập",
        "risk_level": "Medium",
        "test_scenario": "Đã đăng nhập, truy cập lại URL Login — kiểm tra State Transition redirect",
        "pre_condition": "Đã đăng nhập thành công; Đang ở Dashboard",
        "test_steps": (
            "1. Đang ở Dashboard (đã đăng nhập)\n"
            "2. Nhập trực tiếp URL https://crm.anhtester.com/admin/authentication vào thanh địa chỉ\n"
            "3. Nhấn Enter"
        ),
        "test_data": "(Không cần nhập — đã có session)",
        "expected_result": (
            "- Hệ thống redirect về Dashboard /admin/\n"
            "- Không hiển thị lại form Login\n"
            "[A-07]"
        ),
        "priority": "Medium",
    },
    {
        "tc_id": "CRM_LOGIN_TC_015",
        "module": "Đăng nhập",
        "risk_level": "High",
        "test_scenario": "Kiểm tra CSRF token — request thiếu token bị server từ chối (NFR-01)",
        "pre_condition": "Đang ở trang Login",
        "test_steps": (
            "1. Mở trang Login, mở DevTools (F12) → tab Network\n"
            "2. Xác nhận form có trường hidden 'csrf_token_name'\n"
            "3. Dùng Postman gửi POST đến /admin/authentication KHÔNG kèm csrf_token"
        ),
        "test_data": "csrf_token_name: (bị bỏ qua)\nEmail: admin@example.com\nPassword: 12345678",
        "expected_result": (
            "- Server từ chối request (HTTP 403 hoặc lỗi tương đương)\n"
            "- Không đăng nhập được\n"
            "(NFR-01)"
        ),
        "priority": "High",
    },
    # ── MODULE II: FORGOT PASSWORD (MEDIUM RISK) ──────────────────────────────────
    {
        "tc_id": "CRM_LOGIN_TC_016",
        "module": "Forgot Password",
        "risk_level": "Medium",
        "test_scenario": "Click 'Forgot Password?' — chuyển đến đúng URL",
        "pre_condition": "Đang ở trang Login",
        "test_steps": (
            "1. Mở trang Login\n"
            "2. Click vào link 'Forgot Password?'"
        ),
        "test_data": "(Không cần nhập data)",
        "expected_result": (
            "- Trình duyệt chuyển hướng đến:\n"
            "  https://crm.anhtester.com/admin/authentication/forgot_password\n"
            "- Trang hiển thị form với trường Email và nút 'Confirm'"
        ),
        "priority": "Medium",
    },
    {
        "tc_id": "CRM_LOGIN_TC_017",
        "module": "Forgot Password",
        "risk_level": "High",
        "test_scenario": "Nhập email đã đăng ký + Confirm — hệ thống gửi email khôi phục thành công",
        "pre_condition": "Đang ở trang Forgot Password; Email admin@example.com tồn tại trong hệ thống",
        "test_steps": (
            "1. Truy cập /admin/authentication/forgot_password\n"
            "2. Nhập Email đã đăng ký\n"
            "3. Click 'Confirm'"
        ),
        "test_data": "Email: admin@example.com",
        "expected_result": (
            "- Hệ thống gửi email hướng dẫn đặt lại mật khẩu\n"
            "- Hiển thị thông báo xác nhận trên trang\n"
            "- Kiểm tra hộp thư nhận được email hướng dẫn"
        ),
        "priority": "High",
    },
    {
        "tc_id": "CRM_LOGIN_TC_018",
        "module": "Forgot Password",
        "risk_level": "High",
        "test_scenario": "Để trống Email + Click Confirm — HTML5 required validation",
        "pre_condition": "Đang ở trang Forgot Password",
        "test_steps": (
            "1. Truy cập /admin/authentication/forgot_password\n"
            "2. KHÔNG nhập gì vào Email\n"
            "3. Click 'Confirm'"
        ),
        "test_data": "Email: (để trống)",
        "expected_result": (
            "- Trình duyệt chặn form (HTML5 client-side required)\n"
            "- Hiển thị tooltip: 'Please fill out this field.'\n"
            "- Không submit lên server"
        ),
        "priority": "High",
    },
    {
        "tc_id": "CRM_LOGIN_TC_019",
        "module": "Forgot Password",
        "risk_level": "High",
        "test_scenario": "Nhập email không tồn tại + Confirm — kiểm tra bảo mật thông tin (EP)",
        "pre_condition": "Đang ở trang Forgot Password",
        "test_steps": (
            "1. Truy cập /admin/authentication/forgot_password\n"
            "2. Nhập Email đúng định dạng nhưng KHÔNG tồn tại trong hệ thống\n"
            "3. Click 'Confirm'"
        ),
        "test_data": "Email: nonexistent_user_xyz@gmail.com",
        "expected_result": (
            "[A-06] Hệ thống KHÔNG hiển thị 'Email không tồn tại'\n"
            "- Hiển thị thông báo chung (VD: 'Nếu email tồn tại, chúng tôi đã gửi hướng dẫn')\n"
            "- Không tiết lộ thông tin tài khoản (bảo mật)"
        ),
        "priority": "High",
    },
    # ── MODULE III: LOGOUT (LOW RISK) ─────────────────────────────────────────────
    {
        "tc_id": "CRM_LOGIN_TC_020",
        "module": "Đăng xuất",
        "risk_level": "Low",
        "test_scenario": "Đăng xuất thành công — chuyển về trang Login (State Transition)",
        "pre_condition": "Đã đăng nhập; Đang ở Dashboard",
        "test_steps": (
            "1. Từ Dashboard, click vào avatar/profile dropdown (góc trên phải)\n"
            "2. Click 'Logout' trong dropdown"
        ),
        "test_data": "(Không cần nhập data)",
        "expected_result": (
            "- Phiên đăng nhập bị hủy\n"
            "- Chuyển hướng về https://crm.anhtester.com/admin/authentication\n"
            "- Form Login hiển thị lại"
        ),
        "priority": "High",
    },
    {
        "tc_id": "CRM_LOGIN_TC_021",
        "module": "Đăng xuất",
        "risk_level": "Low",
        "test_scenario": "Sau logout, nhấn Back browser — không truy cập được Dashboard (State Transition)",
        "pre_condition": "Đã logout xong (TC_020); Đang ở trang Login",
        "test_steps": (
            "1. Sau khi logout thành công\n"
            "2. Nhấn nút Back của trình duyệt"
        ),
        "test_data": "(Không cần nhập data)",
        "expected_result": (
            "- Hệ thống KHÔNG cho phép quay lại Dashboard\n"
            "- Redirect về trang Login hoặc hiển thị yêu cầu đăng nhập lại\n"
            "- Session cookie đã bị xóa/vô hiệu"
        ),
        "priority": "High",
    },
]

# ─────────────────────────────────────────────
# STYLE CONFIG
# ─────────────────────────────────────────────
COLOR = {
    "header_bg":    "1A237E",   # Xanh navy đậm
    "header_font":  "FFFFFF",   # Trắng
    "high_bg":      "FFCDD2",   # Đỏ nhạt
    "high_font":    "B71C1C",   # Đỏ đậm
    "medium_bg":    "FFF9C4",   # Vàng nhạt
    "medium_font":  "F57F17",   # Vàng đậm
    "low_bg":       "C8E6C9",   # Xanh lá nhạt
    "low_font":     "1B5E20",   # Xanh lá đậm
    "login_bg":     "E3F2FD",   # Xanh dương nhạt (row Login)
    "forgot_bg":    "F3E5F5",   # Tím nhạt (row Forgot Password)
    "logout_bg":    "E8F5E9",   # Xanh lá nhạt (row Logout)
    "title_bg":     "283593",   # Xanh navy tiêu đề section
    "title_font":   "FFFFFF",
    "border":       "90A4AE",   # Xám xanh border
    "alt_row":      "F5F5F5",   # Xám rất nhạt cho alternate row
}

MODULE_ROW_BG = {
    "Đăng nhập":      "EFF8FF",
    "Forgot Password": "FDF5FF",
    "Đăng xuất":       "F0FFF4",
}

def make_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def make_font(hex_color, bold=False, size=11, name="Calibri"):
    return Font(color=hex_color, bold=bold, size=size, name=name)

def make_border():
    s = Side(style="thin", color=COLOR["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def make_align(wrap=True, valign="top", halign="left"):
    return Alignment(wrap_text=wrap, vertical=valign, horizontal=halign)


# ─────────────────────────────────────────────
# BUILD WORKBOOK
# ─────────────────────────────────────────────
wb = openpyxl.Workbook()

# ══════════════════════════════════════════════
# SHEET 1: TEST CASES
# ══════════════════════════════════════════════
ws = wb.active
ws.title = "Test Cases"

# ── Tiêu đề dự án ─────────────────────────────
ws.merge_cells("A1:I1")
cell_title = ws["A1"]
cell_title.value = "📋 BỘ TEST CASE — MODULE LOGIN | Perfex CRM - crm.anhtester.com"
cell_title.font = Font(name="Calibri", bold=True, size=14, color=COLOR["header_font"])
cell_title.fill = make_fill(COLOR["header_bg"])
cell_title.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 35

ws.merge_cells("A2:I2")
cell_sub = ws["A2"]
cell_sub.value = (
    f"Phương pháp: Risk-Based Testing (RBT) | "
    f"Kỹ thuật: EP + BVA + Decision Table + State Transition | "
    f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
)
cell_sub.font = Font(name="Calibri", italic=True, size=10, color="455A64")
cell_sub.fill = make_fill("E8EAF6")
cell_sub.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 20

# ── Header row ────────────────────────────────
HEADERS = [
    "TC ID", "Module", "Risk Level", "Test Scenario",
    "Pre-Condition", "Test Steps", "Test Data",
    "Expected Result", "Priority"
]
COL_WIDTHS = [22, 16, 12, 42, 32, 48, 36, 48, 10]

for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
    cell = ws.cell(row=3, column=col_idx, value=header)
    cell.font = make_font(COLOR["header_font"], bold=True, size=11)
    cell.fill = make_fill(COLOR["header_bg"])
    cell.alignment = make_align(valign="center", halign="center")
    cell.border = make_border()
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[3].height = 28

# ── Data rows ─────────────────────────────────
current_module = None
row_num = 4

for tc in TESTCASES:
    risk = tc["risk_level"]
    module = tc["module"]

    # Màu nền row theo module
    row_bg = MODULE_ROW_BG.get(module, "FFFFFF")

    row_data = [
        tc["tc_id"],
        tc["module"],
        tc["risk_level"],
        tc["test_scenario"],
        tc["pre_condition"],
        tc["test_steps"],
        tc["test_data"],
        tc["expected_result"],
        tc["priority"],
    ]

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.border = make_border()
        cell.alignment = make_align()

        # Màu mặc định theo module
        cell.fill = make_fill(row_bg)
        cell.font = make_font("212121", size=10)

        # Override màu cột Risk Level
        if col_idx == 3:  # Risk Level
            if risk == "High":
                cell.fill = make_fill(COLOR["high_bg"])
                cell.font = make_font(COLOR["high_font"], bold=True, size=10)
                cell.alignment = make_align(valign="center", halign="center")
            elif risk == "Medium":
                cell.fill = make_fill(COLOR["medium_bg"])
                cell.font = make_font(COLOR["medium_font"], bold=True, size=10)
                cell.alignment = make_align(valign="center", halign="center")
            else:
                cell.fill = make_fill(COLOR["low_bg"])
                cell.font = make_font(COLOR["low_font"], bold=True, size=10)
                cell.alignment = make_align(valign="center", halign="center")

        # Override màu cột Priority
        if col_idx == 9:  # Priority
            if tc["priority"] == "High":
                cell.fill = make_fill(COLOR["high_bg"])
                cell.font = make_font(COLOR["high_font"], bold=True, size=10)
            elif tc["priority"] == "Medium":
                cell.fill = make_fill(COLOR["medium_bg"])
                cell.font = make_font(COLOR["medium_font"], bold=True, size=10)
            else:
                cell.fill = make_fill(COLOR["low_bg"])
                cell.font = make_font(COLOR["low_font"], bold=True, size=10)
            cell.alignment = make_align(valign="center", halign="center")

        # TC ID bold
        if col_idx == 1:
            cell.font = make_font("1A237E", bold=True, size=10)

    ws.row_dimensions[row_num].height = 90
    row_num += 1

# Freeze panes (giữ header khi scroll)
ws.freeze_panes = "A4"

# Auto filter
ws.auto_filter.ref = f"A3:I{row_num - 1}"


# ══════════════════════════════════════════════
# SHEET 2: SUMMARY
# ══════════════════════════════════════════════
ws2 = wb.create_sheet("Summary")

summary_title = ws2["A1"]
ws2.merge_cells("A1:D1")
summary_title.value = "📊 TỔNG KẾT BỘ TEST CASE — MODULE LOGIN"
summary_title.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
summary_title.fill = make_fill(COLOR["header_bg"])
summary_title.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

summary_data = [
    ("TỔNG QUAN", "", "", ""),
    ("Tổng số Test Case", "21", "", ""),
    ("🔴 High Priority", "13 TCs", "", ""),
    ("🟡 Medium Priority", "6 TCs", "", ""),
    ("🟢 Low Priority", "2 TCs", "", ""),
    ("", "", "", ""),
    ("PHÂN BỔ THEO MODULE", "", "", ""),
    ("Đăng nhập (HIGH Risk)", "15 TCs", "TC_001 → TC_015", ""),
    ("Forgot Password (MEDIUM Risk)", "4 TCs", "TC_016 → TC_019", ""),
    ("Đăng xuất (LOW Risk)", "2 TCs", "TC_020 → TC_021", ""),
    ("", "", "", ""),
    ("PHÂN BỔ THEO LUỒNG", "", "", ""),
    ("Happy Path", "4 TCs", "TC_001, TC_002, TC_010, TC_017", ""),
    ("Exception Path", "13 TCs", "TC_003~009, TC_011~013, TC_015, TC_018, TC_019", ""),
    ("Alternate Path", "4 TCs", "TC_010, TC_013, TC_014, TC_021", ""),
    ("", "", "", ""),
    ("KỸ THUẬT THIẾT KẾ", "", "", ""),
    ("EP (Equivalence Partitioning)", "TC_001 → TC_009", "", ""),
    ("BVA (Boundary Value Analysis)", "TC_007, TC_011, TC_012, TC_013", "", ""),
    ("Decision Table", "TC_003, TC_004, TC_005", "", ""),
    ("State Transition", "TC_014, TC_020, TC_021", "", ""),
    ("", "", "", ""),
    ("GIẢ ĐỊNH CẦN XÁC NHẬN", "", "", ""),
    ("[A-01] Account Lockout", "Chưa xác định", "Q-01", "🔴 Ưu tiên cao"),
    ("[A-04] Remember me duration", "Giả định 30 ngày", "Q-03", "🟡 Trung bình"),
    ("[A-06] Forgot PW email không tồn tại", "Thông báo chung", "Q-07", "🔴 Ưu tiên cao"),
    ("[A-07] Đã login → truy cập Login URL", "Redirect Dashboard", "Q-08", "🟡 Trung bình"),
]

SECTION_ROWS = {0, 6, 11, 16, 22}

for row_idx, row_vals in enumerate(summary_data, start=2):
    for col_idx, val in enumerate(row_vals, start=1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.border = make_border()
        cell.alignment = make_align(valign="center")
        cell.font = make_font("212121", size=10)
        cell.fill = make_fill("FFFFFF")

        if (row_idx - 2) in SECTION_ROWS and col_idx == 1 and val:
            cell.fill = make_fill(COLOR["title_bg"])
            cell.font = make_font(COLOR["title_font"], bold=True, size=11)
        elif col_idx == 1 and val and (row_idx - 2) not in SECTION_ROWS:
            cell.font = make_font("1A237E", bold=True, size=10)

ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 30
ws2.column_dimensions["C"].width = 45
ws2.column_dimensions["D"].width = 20

# ── Lưu file ──────────────────────────────────
OUTPUT_PATH = "/Users/rachelnguyen/Downloads/Antigravity/antigrativy_Binhtester_kit/testcases_login.xlsx"
wb.save(OUTPUT_PATH)
print(f"\n✅ File đã được xuất thành công!")
print(f"📁 Đường dẫn: {OUTPUT_PATH}")
print(f"📋 Tổng số TC: {len(TESTCASES)}")
print(f"📊 Sheet: 'Test Cases' + 'Summary'")
